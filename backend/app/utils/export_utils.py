import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from app.schemas.report import StudentAttendanceSummary, ReportFilters

TABLE_HEADERS = [
    "Roll No.", "Student Name", "Total Sessions", "Present", "Absent", "Late", "Attendance %",
]


def _rows_from_summaries(students: list[StudentAttendanceSummary]) -> list[list]:
    rows = [TABLE_HEADERS]
    for s in students:
        rows.append(
            [
                s.roll_number,
                s.student_name,
                s.total_sessions,
                s.present_count,
                s.absent_count,
                s.late_count,
                f"{s.attendance_percentage:.1f}%",
            ]
        )
    return rows


def _filters_description(filters: ReportFilters) -> str:
    parts = []
    if filters.date_from:
        parts.append(f"From: {filters.date_from.isoformat()}")
    if filters.date_to:
        parts.append(f"To: {filters.date_to.isoformat()}")
    if filters.class_id:
        parts.append(f"Class ID: {filters.class_id}")
    if filters.subject_id:
        parts.append(f"Subject ID: {filters.subject_id}")
    return " | ".join(parts) if parts else "All records"


def generate_csv_report(students: list[StudentAttendanceSummary]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    for row in _rows_from_summaries(students):
        writer.writerow(row)
    return output.getvalue().encode("utf-8")


def generate_excel_report(students: list[StudentAttendanceSummary], filters: ReportFilters) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    ws.append(["Attendance Report"])
    ws.append([_filters_description(filters)])
    ws.append([])

    header_row_idx = 4
    for row in _rows_from_summaries(students):
        ws.append(row)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[header_row_idx]:
        cell.fill = header_fill
        cell.font = header_font

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        column_letter = column_cells[0].column_letter
        ws.column_dimensions[column_letter].width = max(12, max_length + 2)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_pdf_report(students: list[StudentAttendanceSummary], filters: ReportFilters) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Attendance Report", styles["Title"]))
    elements.append(Paragraph(_filters_description(filters), styles["Normal"]))
    elements.append(Spacer(1, 16))

    data = _rows_from_summaries(students)
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
                ("ALIGN", (2, 0), (-1, -1), "CENTER"),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    return buffer.getvalue()